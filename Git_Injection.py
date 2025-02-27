import os
import logging
import time
from io import BytesIO

import streamlit as st
import pandas as pd
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# =============================
# CONFIGURATION DU LOGGING
# =============================
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# =============================
# CONFIGURATION DE LA PAGE
# =============================
LOGO_PATH = "Centrale-Danone-Logo.png"  # Nom du fichier logo dans le repo
st.set_page_config(page_title="Sythèse de Productivité - Centrale Danone", layout="wide")

st.markdown(
    """
    <style>
    .centered { display: block; margin-left: auto; margin-right: auto; }
    .title { text-align: center; font-size: 50px; font-weight: bold; }
    .subtitle { text-align: center; font-size: 20px; }
    .prestataire { text-align: center; font-size: 30px; font-weight: bold; margin-top: 10px; }
    </style>
    """,
    unsafe_allow_html=True
)

col1, col2, col3 = st.columns([1.5, 2, 1.5])
with col2:
    if os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, width=650, output_format="PNG", caption="", use_column_width=False)
    else:
        st.write("Logo non trouvé.")

st.markdown("<h1 class='title'>Dashboard Productivité - Centrale Danone</h1>", unsafe_allow_html=True)

# =============================
# FONCTION : fig_to_png_bytes
# =============================
def fig_to_png_bytes(fig):
    """Convertit la figure Plotly en PNG et renvoie un objet BytesIO."""
    try:
        img_bytes = fig.to_image(format="png", width=1900, height=900, scale=2)
        return BytesIO(img_bytes)
    except Exception as e:
        logging.error("Erreur lors de la conversion de la figure en PNG : %s", e)
        return None

# =============================
# FONCTION : process_column
# =============================
def process_column(df, column_name):
    """Traite une colonne donnée et retourne un DataFrame formaté."""
    if column_name not in df.columns:
        logging.warning("La colonne '%s' n'existe pas dans le fichier source.", column_name)
        return None

    def categoriser_valeur(x):
        if x < 4000:
            return "<4000"
        elif x <= 8000:
            return "4001-8000"
        elif x <= 11000:
            return "8001-11000"
        elif x <= 14000:
            return "11001-14000"
        else:
            return ">14000"

    df["Tranche"] = df[column_name].apply(categoriser_valeur)
    grouped = df.groupby("Transport", as_index=False).agg({
        column_name: "sum",
        "Matricule": "count"
    })
    grouped.rename(columns={"Matricule": "Nbre SE"}, inplace=True)
    main_col_renamed = column_name

    grouped["KM/SR"] = (grouped[main_col_renamed] / grouped["Nbre SE"]).round(0).astype(int)

    # Comptage par tranche
    count_tranche = df.groupby(["Transport", "Tranche"])["Matricule"].count().reset_index()
    count_tranche = count_tranche.pivot("Transport", "Tranche", "Matricule").fillna(0).reset_index()
    all_tranches = ["<4000", "4001-8000", "8001-11000", "11001-14000", ">14000"]
    for t in all_tranches:
        if t not in count_tranche.columns:
            count_tranche[t] = 0
    count_tranche["Total"] = count_tranche[all_tranches].sum(axis=1)
    for t in all_tranches:
        count_tranche[t + "_%"] = 100 * count_tranche[t] / count_tranche["Total"]

    final = pd.merge(grouped, count_tranche, on="Transport", how="left")
    final.rename(columns={column_name: main_col_renamed}, inplace=True)
    rename_map = {
        "<4000_%": "Taux de réalisation <4000",
        "4001-8000_%": "Taux de réalisation 4001-8000",
        "8001-11000_%": "Taux de réalisation 8001-11000",
        "11001-14000_%": "Taux de réalisation 11001-14000",
        ">14000_%": "Taux de réalisation >14000"
    }
    final.rename(columns=rename_map, inplace=True)

    columns_order = [
        "Transport",
        main_col_renamed,
        "Nbre SE",
        "KM/SR",
        "<4000", "Taux de réalisation <4000",
        "4001-8000", "Taux de réalisation 4001-8000",
        "8001-11000", "Taux de réalisation 8001-11000",
        "11001-14000", "Taux de réalisation 11001-14000",
        ">14000", "Taux de réalisation >14000"
    ]
    final_columns = [c for c in columns_order if c in final.columns]
    final = final[final_columns]

    # Ligne "Total général"
    total_line = {"Transport": "Total général"}
    total_line[main_col_renamed] = final[main_col_renamed].sum()
    total_line["Nbre SE"] = final["Nbre SE"].sum()
    if total_line["Nbre SE"] != 0:
        total_line["KM/SR"] = total_line[main_col_renamed] / total_line["Nbre SE"]
    else:
        total_line["KM/SR"] = 0

    sum_tranches = 0
    for t_col in ["<4000", "4001-8000", "8001-11000", "11001-14000", ">14000"]:
        if t_col in final.columns:
            val_sum = final[t_col].sum()
            total_line[t_col] = val_sum
            sum_tranches += val_sum

    taux_map = {
        "Taux de réalisation <4000": "<4000",
        "Taux de réalisation 4001-8000": "4001-8000",
        "Taux de réalisation 8001-11000": "8001-11000",
        "Taux de réalisation 11001-14000": "11001-14000",
        "Taux de réalisation >14000": ">14000"
    }
    for taux_col, tranche_col in taux_map.items():
        if taux_col in final.columns and tranche_col in total_line:
            if sum_tranches != 0:
                total_line[taux_col] = 100.0 * total_line[tranche_col] / sum_tranches
            else:
                total_line[taux_col] = 0

    row_total = [total_line.get(col, 0) for col in final_columns]
    df_total_line = pd.DataFrame([row_total], columns=final_columns)
    final = pd.concat([final, df_total_line], ignore_index=True)

    # Conversions en int
    final[main_col_renamed] = final[main_col_renamed].astype(int)
    final["Nbre SE"] = final["Nbre SE"].astype(int)
    for t_col in ["<4000", "4001-8000", "8001-11000", "11001-14000", ">14000"]:
        final[t_col] = final[t_col].astype(int)
    for col in final.columns:
        if col.startswith("Taux de réalisation"):
            final[col] = final[col].round(0).astype(int)
    final["KM/SR"] = final["KM/SR"].round(0).astype(int)

    logging.info("Traitement terminé pour la colonne %s.", column_name)
    return final

# =============================
# FONCTION : create_global_file
# =============================
def create_global_file(df_source):
    """
    Traite le DataFrame source et génère un fichier Excel en mémoire (BytesIO)
    avec plusieurs onglets (DRDIST, Atterrissage, Injection).
    Retourne le BytesIO et la liste des onglets créés.
    """
    start_time = time.time()
    logging.info("Début de la création du fichier global.")

    results = {}
    if "Somme de DRDIST" in df_source.columns:
        results["Réel"] = process_column(df_source.copy(), "Somme de DRDIST")
    if "Atterrissage" in df_source.columns:
        results["Atterrissage"] = process_column(df_source.copy(), "Atterrissage")
    if "Injection" in df_source.columns:
        results["Injection"] = process_column(df_source.copy(), "Injection")

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df_result in results.items():
            if df_result is not None:
                df_result.to_excel(writer, sheet_name=sheet_name, index=False)
                logging.info("Écriture de l'onglet %s terminée.", sheet_name)
    output.seek(0)

    wb = load_workbook(output)
    color_header = "0070C0"
    color_header_taux = "002060"
    color_presta = "C0E6F5"
    color_white = "FFFFFF"

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 1) En-tête
        header_row = ws[1]
        for cell in header_row:
            if "Taux de réalisation" in str(cell.value):
                cell.fill = PatternFill(start_color=color_header_taux, end_color=color_header_taux, fill_type="solid")
                cell.font = Font(color=color_white, bold=True)
            else:
                cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
                cell.font = Font(color=color_white, bold=True)

        # 2) Dernière ligne (Total général)
        last_row_index = ws.max_row
        for col_idx in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col_idx).value
            cell = ws.cell(row=last_row_index, column=col_idx)
            if header_value and "Taux de réalisation" in str(header_value):
                cell.fill = PatternFill(start_color=color_header_taux, end_color=color_header_taux, fill_type="solid")
                cell.font = Font(color=color_white, bold=True)
            else:
                cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
                cell.font = Font(color=color_white, bold=True)

        # 3) Colonne "Transport" pour les lignes de données
        for row_idx in range(2, last_row_index):
            cell = ws.cell(row=row_idx, column=1)
            cell.fill = PatternFill(start_color=color_presta, end_color=color_presta, fill_type="solid")

        # 4) Ajouter le symbole "%" dans les colonnes de taux
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                header_value = ws.cell(row=1, column=col_idx).value
                if header_value and "Taux de réalisation" in str(header_value):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    ws.cell(row=row_idx, column=col_idx).value = f"{val}%"

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    elapsed = time.time() - start_time
    logging.info("Création du fichier global terminée en %.2f secondes.", elapsed)
    return final_output, list(results.keys())

# =============================
# FONCTION : make_plots_for_sheet
# =============================
def make_plots_for_sheet(sheet_name, file_excel):
    logging.info("Génération des graphiques pour l'onglet %s.", sheet_name)
    df = pd.read_excel(file_excel, sheet_name=sheet_name)

    # Colonnes de taux (on retire '%' et on convertit en float)
    cols_taux = [
        "Taux de réalisation <4000",
        "Taux de réalisation 4001-8000",
        "Taux de réalisation 8001-11000",
        "Taux de réalisation 11001-14000",
        "Taux de réalisation >14000"
    ]
    for col in cols_taux:
        df[col] = df[col].astype(str).str.replace('%', '', regex=False).astype(float)

    # Colonnes de nombre
    cols_nb = ["<4000", "4001-8000", "8001-11000", "11001-14000", ">14000"]

    df_total = df[df["Transport"] == "Total général"].copy()
    df_presta = df[df["Transport"] != "Total général"].copy()

    # FIG1
    df_taux_total = pd.melt(
        df_total,
        id_vars=["Transport"],
        value_vars=cols_taux,
        var_name="Palier",
        value_name="Taux (%)"
    )
    df_taux_presta["Palier"] = df_taux_presta["Palier"].str.replace("Taux de réalisation ", "", regex=False)

    fig1 = px.bar(
        df_taux_total,
        x="Palier",
        y="Taux (%)",
        text="Taux (%)",
        title=f"Taux de réalisation par palier ({sheet_name}) ",
        color_discrete_sequence=["#1f4486"]
    )
    fig1.update_layout(template=None)
    fig1.update_layout(
        title=dict(
            text=fig1.layout.title.text,
            x=0.5,
            xanchor='center',
            font=dict(size=22, family="Arial Black", color="black")
        ),
        paper_bgcolor="white",
        plot_bgcolor="white",
        font=dict(color="black"),
        xaxis=dict(
            title="Palier Kilométrique",
            color="black",
            showline=True,
            linecolor="black",
            tickfont=dict(color="black"),
            title_font=dict(size=16, family="Arial Black", color="black")
        ),
        yaxis=dict(
            title="Taux De Réalisation (%)",
            color="black",
            showline=True,
            linecolor="black",
            tickfont=dict(color="black"),
            title_font=dict(size=16, family="Arial Black", color="black")
        ),
        legend=dict(
            title_font=dict(color="black", size=14, family="Arial Black"),
            font=dict(color="black", size=12)
        )
    )
    fig1.update_yaxes(
        tickmode='array',
        tickvals=[0, 25, 50, 75, 100],
        ticktext=['0%', '25%', '50%', '75%', '100%'],
        range=[0, 100],
        showgrid=False
    )
    fig1.update_traces(
        texttemplate='%{y}%',
        textposition='outside',
        textfont=dict(color='black')
    )

    # FIG2
    df_taux_presta = pd.melt(
        df_presta,
        id_vars=["Transport"],
        value_vars=cols_taux,
        var_name="Palier",
        value_name="Taux (%)"
    )
    df_taux_presta["Palier"] = df_taux_presta["Palier"].str.replace("Taux de réalisation ", "", regex=False)

    fig2 = px.bar(
        df_taux_presta,
        x="Transport",
        y="Taux (%)",
        text="Taux (%)",
        color="Palier",
        barmode="group",
        title=f"Taux de réalisation par palier et par prestataire ({sheet_name})",
        color_discrete_sequence=["#1f4486", "#3d9ddb", "#dd2b17", "#275c20", "#dca433"]
    )
    fig2.update_layout(template=None)
    fig2.update_layout(
        title=dict(
            text=fig2.layout.title.text,
            x=0.5,
            xanchor='center',
            font=dict(size=22, family="Arial Black", color="black")
        ),
        paper_bgcolor="white",
        plot_bgcolor="white",
        font=dict(color="black"),
        xaxis=dict(
            title="Prestataire",
            color="black",
            showline=True,
            linecolor="black",
            tickfont=dict(color="black"),
            title_font=dict(size=16, family="Arial Black", color="black")
        ),
        yaxis=dict(
            title="Taux De Réalisation (%)",
            color="black",
            showline=True,
            linecolor="black",
            tickfont=dict(color="black"),
            title_font=dict(size=16, family="Arial Black", color="black")
        ),
        legend=dict(
            title_font=dict(color="black", size=14, family="Arial Black"),
            font=dict(color="black", size=12)
        )
    )
    fig2.update_yaxes(
        tickmode='array',
        tickvals=[0, 25, 50, 75, 100],
        ticktext=['0%', '25%', '50%', '75%', '100%'],
        range=[0, 100],
        showgrid=False
    )
    fig2.update_traces(
        texttemplate='%{y}%',
        textposition='outside',
        textfont=dict(color='black')
    )

    # FIG3
    df_nb_presta = pd.melt(
        df_presta,
        id_vars=["Transport"],
        value_vars=cols_nb,
        var_name="Palier",
        value_name="Nombre"
    )
    df_nb_presta["Palier"] = df_nb_presta["Palier"].apply(lambda x: f"({x})")

    fig3 = px.bar(
        df_nb_presta,
        x="Transport",
        y="Nombre",
        text="Nombre",
        color="Palier",
        barmode="group",
        title=f"Nombre d'éléments par palier et par prestataire ({sheet_name})",
        color_discrete_sequence=["#1f4486", "#3d9ddb", "#dd2b17", "#275c20", "#dca433"]
    )
    fig3.update_layout(template=None)
    fig3.update_layout(
        title=dict(
            text=fig3.layout.title.text,
            x=0.5,
            xanchor='center',
            font=dict(size=22, family="Arial Black", color="black")
        ),
        paper_bgcolor="white",
        plot_bgcolor="white",
        font=dict(color="black"),
        xaxis=dict(
            title="Prestataire",
            color="black",
            showline=True,
            linecolor="black",
            tickfont=dict(color="black"),
            title_font=dict(size=16, family="Arial Black", color="black")
        ),
        yaxis=dict(
            title="Nombre",
            color="black",
            showline=True,
            linecolor="black",
            tickfont=dict(color="black"),
            title_font=dict(size=16, family="Arial Black", color="black")
        ),
        legend=dict(
            title_font=dict(color="black", size=14, family="Arial Black"),
            font=dict(color="black", size=12)
        )
    )
    fig3.update_xaxes(showgrid=False)
    fig3.update_yaxes(showgrid=False)
    fig3.update_traces(
        texttemplate='%{y}',
        textposition='outside',
        textfont=dict(color='black')
    )

    return fig1, fig2, fig3

# =============================
# FONCTION PRINCIPALE STREAMLIT
# =============================
def main():
    st.title("Analyse Recap - DRDIST / Atterrissage / Injection")
    uploaded_file = st.file_uploader("Choisissez votre fichier Excel (Recap.xlsx)", type=["xlsx"])

    if uploaded_file is not None:
        df_source = pd.read_excel(uploaded_file)
        logging.info("Fichier %s chargé.", uploaded_file.name)
        st.write("Fichier chargé avec succès. Traitement en cours...")

        # Création du fichier global en mémoire
        file_global, sheet_list = create_global_file(df_source)
        st.success("Le fichier_resultat_global.xlsx a été généré en mémoire.")
        logging.info("Onglets générés : %s", sheet_list)

        # Affichage des graphiques pour chaque onglet
        for sheet_name in sheet_list:
            st.subheader(f"Visualisation pour l'onglet : {sheet_name}")
            fig1, fig2, fig3 = make_plots_for_sheet(sheet_name, file_global)

            # Graphique 1
            st.plotly_chart(fig1, use_container_width=True)
            png_bytes1 = fig_to_png_bytes(fig1)
            if png_bytes1:
                st.download_button(
                    label=f"Télécharger {sheet_name} - Graphique 1 (PNG)",
                    data=png_bytes1,
                    file_name=f"{sheet_name}_graph1.png",
                    mime="image/png"
                )

            # Graphique 2
            st.plotly_chart(fig2, use_container_width=True)
            png_bytes2 = fig_to_png_bytes(fig2)
            if png_bytes2:
                st.download_button(
                    label=f"Télécharger {sheet_name} - Graphique 2 (PNG)",
                    data=png_bytes2,
                    file_name=f"{sheet_name}_graph2.png",
                    mime="image/png"
                )

            # Graphique 3
            st.plotly_chart(fig3, use_container_width=True)
            png_bytes3 = fig_to_png_bytes(fig3)
            if png_bytes3:
                st.download_button(
                    label=f"Télécharger {sheet_name} - Graphique 3 (PNG)",
                    data=png_bytes3,
                    file_name=f"{sheet_name}_graph3.png",
                    mime="image/png"
                )

        # Bouton pour télécharger le fichier Excel final
        st.download_button(
            label="Télécharger le fichier_resultat_global.xlsx",
            data=file_global,
            file_name="fichier_resultat_global.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("Veuillez charger un fichier Excel pour commencer.")

if __name__ == "__main__":
    main()
